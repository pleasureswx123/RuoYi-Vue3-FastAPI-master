import io

import pytest
from openpyxl import Workbook
from pydantic import ValidationError

from module_shot_grid.config import ShotGridImportConfig
from module_shot_grid.entity.vo.asset_import_vo import AssetImportCommitRequestModel
from module_shot_grid.exceptions import ShotGridDomainException
from module_shot_grid.service.asset_excel_parser import AssetExcelParser
from module_shot_grid.service.asset_import_template_service import ShotGridAssetImportTemplateService

ASSET_SAMPLE = ShotGridAssetImportTemplateService.TEMPLATE_PATH
HEADERS = ['类型', '名称', '描述', '制作分项', '备注', '状态']
EXCEL_LAST_ROW = 1_048_576
DUPLICATE_ROW_COUNT = 2


def _save_workbook(workbook: Workbook) -> bytes:
    stream = io.BytesIO()
    workbook.save(stream)
    workbook.close()
    return stream.getvalue()


def _minimal_workbook() -> Workbook:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Sheet1'
    sheet.append(HEADERS)
    sheet.append(['场景', '控制室', '控制室描述', '恐怖气氛主视角', '重点', None])
    return workbook


def test_real_asset_sample_has_frozen_counts_and_warnings() -> None:
    result = AssetExcelParser().parse(ASSET_SAMPLE.read_bytes())

    assert result.summary.model_dump(by_alias=True) == {
        'totalRows': 6,
        'validRows': 6,
        'warningRows': 0,
        'errorRows': 0,
        'distinctAssets': 4,
        'distinctAssetItems': 6,
        'byType': {
            'Character': {'assets': 1, 'items': 2, 'validRows': 2, 'warningRows': 0, 'errorRows': 0},
            'Environment': {'assets': 1, 'items': 2, 'validRows': 2, 'warningRows': 0, 'errorRows': 0},
            'Prop': {'assets': 2, 'items': 2, 'validRows': 2, 'warningRows': 0, 'errorRows': 0},
        },
        'estimatedAutoMatches': 0,
    }
    assert [warning.error_key for warning in result.workbook_warnings] == ['SG_IMPORT_READONLY_COLUMNS_IGNORED']
    assert not any(row.warnings or row.errors for row in result.rows)


def test_merged_parent_cells_are_inherited_without_blind_forward_fill() -> None:
    result = AssetExcelParser().parse(ASSET_SAMPLE.read_bytes())
    by_row = {row.row_number: row for row in result.rows}

    assert by_row[2].normalized.asset_name == by_row[3].normalized.asset_name
    assert by_row[2].normalized.asset_description
    assert by_row[2].normalized.asset_description == by_row[3].normalized.asset_description
    assert by_row[4].normalized.asset_name == by_row[5].normalized.asset_name
    assert by_row[4].normalized.asset_description
    assert by_row[4].normalized.asset_description == by_row[5].normalized.asset_description
    assert all(row.normalized.item_description is None for row in result.rows)

    workbook = _minimal_workbook()
    workbook.active.append([None, None, None, '第二分项'])
    parsed = AssetExcelParser().parse(_save_workbook(workbook))
    assert parsed.rows[1].normalized.asset_name is None
    assert {'SG_ASSET_TYPE_INVALID', 'SG_ASSET_NAME_REQUIRED'} <= {issue.error_key for issue in parsed.rows[1].errors}


@pytest.mark.parametrize('description_header', ['描述', '资产描述'])
def test_asset_description_is_inherited_but_item_requirements_stay_separate(description_header: str) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(['类型', '名称', description_header, '制作分项', '分项补充要求', '备注'])
    sheet.append(['场景', '测试舱室', '斑驳铁皮墙和狭小舱室', '主视角', '从门口看向睡袋', '主视角备注'])
    sheet.append([None, None, None, '反打视角', '从睡袋看向铁皮门', '反打备注'])
    for column in ('A', 'B', 'C'):
        sheet.merge_cells(f'{column}2:{column}3')
    result = AssetExcelParser().parse(_save_workbook(workbook))
    assert result.summary.error_rows == 0
    assert [row.normalized.asset_description for row in result.rows] == ['斑驳铁皮墙和狭小舱室'] * 2
    assert [row.normalized.item_description for row in result.rows] == ['从门口看向睡袋', '从睡袋看向铁皮门']
    assert [row.normalized.remark for row in result.rows] == ['主视角备注', '反打备注']


def test_different_asset_descriptions_for_one_asset_are_rejected() -> None:
    workbook = _minimal_workbook()
    workbook.active.append(['场景', '控制室', '冲突的资产描述', '反打视角'])
    result = AssetExcelParser().parse(_save_workbook(workbook))
    assert all(any(issue.field_name == 'assetDescription' for issue in row.errors) for row in result.rows)


def test_status_formula_and_columns_after_first_blank_header_are_ignored() -> None:
    workbook = _minimal_workbook()
    sheet = workbook.active
    sheet['F2'] = '=1+1'
    sheet['I1'] = '辅助列'
    sheet['I2'] = '=1+1'

    result = AssetExcelParser().parse(_save_workbook(workbook))

    assert result.summary.total_rows == 1
    assert result.rows[0].errors == []
    assert [warning.error_key for warning in result.workbook_warnings] == ['SG_IMPORT_READONLY_COLUMNS_IGNORED']


def test_hidden_sheets_add_one_workbook_warning() -> None:
    workbook = _minimal_workbook()
    workbook.create_sheet('隐藏1').sheet_state = 'hidden'
    workbook.create_sheet('隐藏2').sheet_state = 'hidden'

    result = AssetExcelParser().parse(_save_workbook(workbook))

    assert [warning.error_key for warning in result.workbook_warnings] == [
        'SG_IMPORT_HIDDEN_SHEETS_IGNORED',
        'SG_IMPORT_READONLY_COLUMNS_IGNORED',
    ]


def test_duplicate_production_item_is_rejected_across_sheets() -> None:
    workbook = _minimal_workbook()
    second = workbook.create_sheet('Sheet2')
    second.append(HEADERS)
    second.append(['Environment', '控制室', '另一描述', '恐怖气氛主视角'])

    result = AssetExcelParser().parse(_save_workbook(workbook))

    assert result.summary.error_rows == DUPLICATE_ROW_COUNT
    assert all(
        any(issue.error_key == 'SG_ASSET_PRODUCTION_ITEM_CONFLICT' for issue in row.errors) for row in result.rows
    )


def test_only_parent_columns_allow_single_column_vertical_merges() -> None:
    workbook = _minimal_workbook()
    sheet = workbook.active
    sheet.append(['场景', '第二资产', '描述', '分项'])
    sheet.merge_cells('B2:C2')

    with pytest.raises(ShotGridDomainException) as exc_info:
        AssetExcelParser().parse(_save_workbook(workbook))

    assert exc_info.value.error_key == 'SG_IMPORT_TEMPLATE_INVALID'


def test_row_limit_rejects_sparse_malicious_tail_before_full_scan() -> None:
    workbook = _minimal_workbook()
    workbook.active.cell(row=EXCEL_LAST_ROW, column=1, value='恶意尾行')

    with pytest.raises(ShotGridDomainException) as exc_info:
        AssetExcelParser(ShotGridImportConfig(max_rows_per_workbook=10)).parse(_save_workbook(workbook))

    assert exc_info.value.error_key == 'SG_IMPORT_ROW_LIMIT_EXCEEDED'


@pytest.mark.parametrize(
    ('column', 'value', 'field_name'),
    [
        ('B', 'İ' * 200, 'assetName'),
        ('D', 'İ' * 240, 'productionItem'),
        ('E', '备' * 501, 'remark'),
    ],
)
def test_values_and_casefold_keys_fit_database_lengths(column: str, value: str, field_name: str) -> None:
    workbook = _minimal_workbook()
    workbook.active[f'{column}2'] = value

    result = AssetExcelParser().parse(_save_workbook(workbook))

    assert any(
        issue.error_key == 'SG_IMPORT_FIELD_TOO_LONG' and issue.field_name == field_name
        for issue in result.rows[0].errors
    )


def test_commit_selection_is_composite_and_allows_partial_parent_group() -> None:
    request = AssetImportCommitRequestModel(
        importToken='token',
        selectedRows=[{'sheetName': 'Sheet1', 'rowNumber': 2}],
    )
    assert [row.key() for row in request.selected_rows] == [('Sheet1', 2)]

    with pytest.raises(ValidationError):
        AssetImportCommitRequestModel(
            importToken='token',
            selectedRows=[{'sheetName': 'Sheet1', 'rowNumber': 2, 'assigneeUserId': 7}],
        )

    with pytest.raises(ValueError):
        AssetImportCommitRequestModel(
            importToken='token',
            selectedRows=[
                {'sheetName': 'Sheet1', 'rowNumber': 2},
                {'sheetName': 'Sheet1', 'rowNumber': 2},
            ],
        )


def test_assignee_column_is_not_supported_by_current_template() -> None:
    workbook = _minimal_workbook()
    workbook.active['G1'] = '制作人'

    with pytest.raises(ShotGridDomainException) as exc_info:
        AssetExcelParser().parse(_save_workbook(workbook))

    assert exc_info.value.error_key == 'SG_IMPORT_HEADER_INVALID'
