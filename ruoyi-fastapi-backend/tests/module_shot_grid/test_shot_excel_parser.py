import io
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from module_shot_grid.config import ShotGridImportConfig
from module_shot_grid.entity.vo.shot_import_vo import ShotImportCommitRequestModel
from module_shot_grid.exceptions import ShotGridDomainException
from module_shot_grid.service.shot_excel_parser import ShotExcelParser

REPO_ROOT = Path(__file__).resolve().parents[3]
SHOT_SAMPLE = REPO_ROOT / 'shot-grid-frontend' / 'docs' / '镜头-样表.xlsx'
FIRST_SORT_ORDER = 10
SQL_INTEGER_OVERFLOW = 2_147_483_648
SQL_BIGINT_MILLISECONDS_OVERFLOW_SECONDS = '9223372036854775.808'


def _save_workbook(workbook: Workbook) -> bytes:
    stream = io.BytesIO()
    workbook.save(stream)
    workbook.close()
    return stream.getvalue()


def _minimal_workbook() -> Workbook:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'EP001'
    sheet.append(list(ShotExcelParser.EXPECTED_HEADERS))
    sheet.append(['序', 'S001', 1.5, None, None, '镜头描述', None, None, None, '35/25', '太空'])
    return workbook


def test_real_sample_has_frozen_structure_and_exact_counts() -> None:
    result = ShotExcelParser().parse(SHOT_SAMPLE.read_bytes())

    assert result.summary.model_dump(by_alias=True) == {
        'totalRows': 24,
        'validRows': 24,
        'warningRows': 0,
        'errorRows': 0,
        'distinctEpisodes': 2,
        'distinctScenes': 8,
        'distinctShots': 24,
    }
    assert [row.sheet_name for row in result.rows[:13]] == ['EP001'] * 12 + ['EP002']
    first = result.rows[0].normalized
    assert first is not None
    assert (first.episode_no, first.episode_code) == (1, 'EP001')
    assert (first.scene_no, first.scene_code, first.scene_name) == (0, '000', '序')
    assert (first.shot_no, first.shot_code, first.duration_ms, first.focal_length) == (1, 'S001', 8000, '135')
    assert first.sort_order == FIRST_SORT_ORDER
    assert first.asset_requirements[0].raw_name == '太空'
    assert result.rows[2].normalized is not None
    assert result.rows[2].normalized.focal_length == '35/25'
    assert result.rows[3].normalized is not None
    assert result.rows[3].normalized.focal_length == '24/18'
    assert [warning.error_key for warning in result.workbook_warnings] == ['SG_IMPORT_READONLY_COLUMNS_IGNORED']


def test_parser_ignores_auxiliary_columns_after_first_blank_header() -> None:
    workbook = _minimal_workbook()
    sheet = workbook.active
    sheet['S1'] = '不应导入'
    sheet['S2'] = '=1+1'

    result = ShotExcelParser().parse(_save_workbook(workbook))

    assert result.summary.valid_rows == 1
    assert result.rows[0].errors == []


def test_parser_ignores_readonly_values_and_formulas() -> None:
    workbook = _minimal_workbook()
    sheet = workbook.active
    sheet['E2'] = '=1+1'
    sheet['P2'] = '=1+1'
    sheet.append([None, None, None, None, '只读缩略图', *([None] * 10), 'completed'])

    result = ShotExcelParser().parse(_save_workbook(workbook))

    assert result.summary.total_rows == 1
    assert result.summary.valid_rows == 1
    assert result.rows[0].errors == []


def test_duplicate_shot_number_across_scenes_is_row_error() -> None:
    workbook = _minimal_workbook()
    sheet = workbook.active
    sheet.append(['01场', 'S001', 1, None, None, '第二个镜头'])

    result = ShotExcelParser().parse(_save_workbook(workbook))

    assert result.summary.valid_rows == 1
    assert result.summary.error_rows == 1
    assert result.rows[1].errors[0].error_key == 'SG_SHOT_NO_CONFLICT'


@pytest.mark.parametrize('duration', ['1.2345', -1, 'abc'])
def test_invalid_duration_is_rejected(duration: object) -> None:
    workbook = _minimal_workbook()
    workbook.active['C2'] = duration

    result = ShotExcelParser().parse(_save_workbook(workbook))

    assert result.rows[0].can_import is False
    assert any(issue.error_key == 'SG_IMPORT_DURATION_INVALID' for issue in result.rows[0].errors)


def test_formula_in_main_region_is_rejected() -> None:
    workbook = _minimal_workbook()
    workbook.active['C2'] = '=1+1'

    result = ShotExcelParser().parse(_save_workbook(workbook))

    assert result.rows[0].can_import is False
    assert any(issue.error_key == 'SG_IMPORT_FORMULA_NOT_ALLOWED' for issue in result.rows[0].errors)


def test_positive_scene_name_is_not_persisted_from_code_cell() -> None:
    workbook = _minimal_workbook()
    workbook.active['A2'] = '01场'

    result = ShotExcelParser().parse(_save_workbook(workbook))

    normalized = result.rows[0].normalized
    assert normalized is not None
    assert (normalized.scene_no, normalized.scene_code, normalized.scene_name) == (1, '001', None)


def test_header_mismatch_is_workbook_error() -> None:
    workbook = _minimal_workbook()
    workbook.active['F1'] = '错误表头'

    with pytest.raises(ShotGridDomainException) as exc_info:
        ShotExcelParser().parse(_save_workbook(workbook))

    assert exc_info.value.error_key == 'SG_IMPORT_HEADER_MISMATCH'


def test_row_limit_uses_frozen_error_key() -> None:
    workbook = _minimal_workbook()
    workbook.active.append(['01场', 'S002', 1, None, None, '第二个镜头'])

    with pytest.raises(ShotGridDomainException) as exc_info:
        ShotExcelParser(ShotGridImportConfig(max_rows_per_workbook=1)).parse(_save_workbook(workbook))

    assert exc_info.value.error_key == 'SG_IMPORT_ROW_LIMIT_EXCEEDED'


def test_row_limit_rejects_large_physical_max_row_before_scanning() -> None:
    workbook = _minimal_workbook()
    workbook.active.cell(row=1_048_576, column=1, value='恶意尾行')

    with pytest.raises(ShotGridDomainException) as exc_info:
        ShotExcelParser(ShotGridImportConfig(max_rows_per_workbook=10)).parse(_save_workbook(workbook))

    assert exc_info.value.error_key == 'SG_IMPORT_ROW_LIMIT_EXCEEDED'


def test_episode_number_must_fit_database_integer() -> None:
    workbook = _minimal_workbook()
    workbook.active.title = f'EP{SQL_INTEGER_OVERFLOW}'

    with pytest.raises(ShotGridDomainException) as exc_info:
        ShotExcelParser().parse(_save_workbook(workbook))

    assert exc_info.value.error_key == 'SG_IMPORT_SHEET_NAME_INVALID'


@pytest.mark.parametrize(
    ('column', 'value', 'error_key'),
    [
        ('A', str(SQL_INTEGER_OVERFLOW), 'SG_IMPORT_SCENE_INVALID'),
        ('B', f'S{SQL_INTEGER_OVERFLOW}', 'SG_IMPORT_SHOT_NO_INVALID'),
        ('C', SQL_BIGINT_MILLISECONDS_OVERFLOW_SECONDS, 'SG_IMPORT_DURATION_INVALID'),
        ('D', '制' * 31, 'SG_IMPORT_FIELD_TOO_LONG'),
    ],
)
def test_row_values_must_fit_database_and_member_boundaries(column: str, value: object, error_key: str) -> None:
    workbook = _minimal_workbook()
    workbook.active[f'{column}2'] = value

    result = ShotExcelParser().parse(_save_workbook(workbook))

    assert result.rows[0].can_import is False
    assert any(issue.error_key == error_key for issue in result.rows[0].errors)


def test_commit_selection_uses_sheet_and_row_composite_identity() -> None:
    request = ShotImportCommitRequestModel(
        importToken='token',
        selectedRows=[
            {'sheetName': 'EP001', 'rowNumber': 2},
            {'sheetName': 'EP002', 'rowNumber': 2},
        ],
    )
    assert [row.key() for row in request.selected_rows] == [('EP001', 2), ('EP002', 2)]

    with pytest.raises(ValueError):
        ShotImportCommitRequestModel(
            importToken='token',
            selectedRows=[
                {'sheetName': 'EP001', 'rowNumber': 2},
                {'sheetName': 'EP001', 'rowNumber': 2},
            ],
        )


def test_real_sample_has_no_formula_in_main_region() -> None:
    workbook = load_workbook(SHOT_SAMPLE, data_only=False, read_only=False)
    try:
        assert not any(
            cell.data_type == 'f'
            for sheet in workbook.worksheets
            for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=16)
            for cell in row
        )
    finally:
        workbook.close()
