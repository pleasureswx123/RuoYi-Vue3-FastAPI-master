import hashlib
from zipfile import ZipFile

import pytest
from openpyxl import load_workbook

from module_shot_grid.controller.asset_import_template_controller import (
    asset_import_template_controller,
    download_asset_import_template,
)
from module_shot_grid.service.asset_excel_parser import AssetExcelParser
from module_shot_grid.service.asset_import_template_service import ShotGridAssetImportTemplateService

HTTP_OK = 200
EXPECTED_ASSET_ROWS = 6
EXPECTED_SHA256 = 'b551ac1d1d5edc20a025b0ed90157412e1365006108816f08cb2c59ae4301696'
EXPECTED_HEADERS = ('类型', '名称', '描述', '制作分项', '备注', '状态')


def test_packaged_asset_template_digest_and_anonymous_content_are_frozen() -> None:
    contents = ShotGridAssetImportTemplateService.TEMPLATE_PATH.read_bytes()

    assert hashlib.sha256(contents).hexdigest() == EXPECTED_SHA256
    assert ShotGridAssetImportTemplateService.EXPECTED_SHA256 == EXPECTED_SHA256
    with ZipFile(ShotGridAssetImportTemplateService.TEMPLATE_PATH) as workbook:
        payload = b'\n'.join(workbook.read(name) for name in workbook.namelist() if name.endswith('.xml'))
    assert b'D:\\' not in payload
    assert b'file:' not in payload.lower()
    assert b'example.com' not in payload.lower()


def test_packaged_asset_template_uses_exact_headers_without_assignee_column() -> None:
    workbook = load_workbook(ShotGridAssetImportTemplateService.TEMPLATE_PATH, read_only=True, data_only=False)
    try:
        visible_sheets = [sheet for sheet in workbook.worksheets if sheet.sheet_state == 'visible']
        assert visible_sheets
        for sheet in visible_sheets:
            assert tuple(sheet.cell(row=1, column=index).value for index in range(1, 7)) == EXPECTED_HEADERS
            assert not any(
                '制作人' in str(cell.value)
                for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row)
                for cell in row
                if cell.value is not None
            )
    finally:
        workbook.close()

    parsed = AssetExcelParser().parse(ShotGridAssetImportTemplateService.TEMPLATE_PATH.read_bytes())
    assert parsed.summary.total_rows == EXPECTED_ASSET_ROWS
    assert parsed.summary.error_rows == 0


@pytest.mark.asyncio
async def test_asset_template_route_returns_versioned_download_headers() -> None:
    response = await download_asset_import_template(None)  # type: ignore[arg-type]

    assert response.status_code == HTTP_OK
    assert response.body == ShotGridAssetImportTemplateService.TEMPLATE_PATH.read_bytes()
    assert response.headers['x-shot-grid-template-version'] == 'asset-v2'
    assert 'asset-import-template-asset-v2.xlsx' in response.headers['content-disposition']


def test_asset_template_route_uses_frozen_path() -> None:
    route = asset_import_template_controller.routes[0]

    assert route.path == '/shot-grid/imports/assets/template'
    assert route.methods == {'GET'}
