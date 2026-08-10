import io
import zipfile
from pathlib import Path

import pytest
from openpyxl import Workbook

from module_shot_grid.config import ShotGridImportConfig
from module_shot_grid.exceptions import ShotGridDomainException
from module_shot_grid.service.excel_security_service import ExcelSecurityService

SHA256_HEX_LENGTH = 64
TEST_CONTENT_LENGTH = 3
PAYLOAD_TOO_LARGE_STATUS = 413
UNPROCESSABLE_STATUS = 422
TEST_PREVIEW_TTL_SECONDS = 60


def _xlsx_bytes() -> bytes:
    workbook = Workbook()
    stream = io.BytesIO()
    workbook.save(stream)
    workbook.close()
    return stream.getvalue()


def _append_zip_entry(contents: bytes, name: str) -> bytes:
    archive_name = name.replace('\\', '/')
    source = zipfile.ZipFile(io.BytesIO(contents))
    output = io.BytesIO()
    with source, zipfile.ZipFile(output, mode='w', compression=zipfile.ZIP_DEFLATED) as target:
        for entry in source.infolist():
            target.writestr(entry, source.read(entry.filename))
        target.writestr(archive_name, b'<xml/>')
    result = output.getvalue()
    if archive_name != name:
        result = result.replace(archive_name.encode(), name.encode())
    return result


def _replace_zip_entries(contents: bytes, replacements: dict[str, bytes]) -> bytes:
    source = zipfile.ZipFile(io.BytesIO(contents))
    output = io.BytesIO()
    with source, zipfile.ZipFile(output, mode='w', compression=zipfile.ZIP_DEFLATED) as target:
        for entry in source.infolist():
            if entry.filename not in replacements:
                target.writestr(entry, source.read(entry.filename))
        for name, value in replacements.items():
            target.writestr(name, value)
    return output.getvalue()


def test_validate_and_hash_accepts_real_xlsx_container() -> None:
    contents = _xlsx_bytes()
    digest = ExcelSecurityService.validate_and_hash('镜头.xlsx', contents)
    assert len(digest) == SHA256_HEX_LENGTH
    assert digest == digest.lower()


@pytest.mark.parametrize('sample_name', ['镜头-样表.xlsx', '资产-样表.xlsx'])
def test_validate_and_hash_keeps_formal_samples_compatible(sample_name: str) -> None:
    sample = Path(__file__).parents[3] / 'shot-grid-frontend' / 'docs' / sample_name

    digest = ExcelSecurityService.validate_and_hash(sample.name, sample.read_bytes())

    assert len(digest) == SHA256_HEX_LENGTH


@pytest.mark.parametrize(
    ('file_name', 'contents', 'error_key'),
    [
        ('镜头.xls', b'not-xlsx', 'SG_IMPORT_FILE_TYPE_INVALID'),
        ('镜头.xlsx', b'', 'SG_IMPORT_FILE_EMPTY'),
        ('镜头.xlsx', b'not-a-zip', 'SG_IMPORT_FILE_INVALID'),
    ],
)
def test_validate_and_hash_rejects_invalid_file(file_name: str, contents: bytes, error_key: str) -> None:
    with pytest.raises(ShotGridDomainException) as exc_info:
        ExcelSecurityService.validate_and_hash(file_name, contents)
    assert exc_info.value.error_key == error_key


def test_validate_and_hash_enforces_binary_size_limit() -> None:
    with pytest.raises(ShotGridDomainException) as exc_info:
        ExcelSecurityService.validate_and_hash(
            '镜头.xlsx',
            _xlsx_bytes(),
            ShotGridImportConfig(max_file_size_bytes=1),
        )
    assert exc_info.value.error_key == 'SG_IMPORT_FILE_TOO_LARGE'
    assert exc_info.value.http_status == PAYLOAD_TOO_LARGE_STATUS


def test_ooxml_preflight_counts_hidden_sheet_rows_before_openpyxl() -> None:
    workbook = Workbook()
    workbook.active['A1'] = '可见'
    hidden = workbook.create_sheet('隐藏')
    hidden.sheet_state = 'hidden'
    hidden['A1'] = '一'
    hidden['A2'] = '二'
    stream = io.BytesIO()
    workbook.save(stream)
    workbook.close()

    with pytest.raises(ShotGridDomainException) as exc_info:
        ExcelSecurityService.validate_and_hash(
            '镜头.xlsx',
            stream.getvalue(),
            ShotGridImportConfig(max_ooxml_rows_per_workbook=2),
        )

    assert exc_info.value.http_status == UNPROCESSABLE_STATUS
    assert exc_info.value.error_key == 'SG_IMPORT_ROW_LIMIT_EXCEEDED'


def test_ooxml_preflight_limits_physical_cells_and_merge_expansion() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(['一', '二', '三'])
    worksheet.merge_cells('A2:A4')
    stream = io.BytesIO()
    workbook.save(stream)
    workbook.close()

    with pytest.raises(ShotGridDomainException) as cell_exc:
        ExcelSecurityService.validate_and_hash(
            '资产.xlsx',
            stream.getvalue(),
            ShotGridImportConfig(max_ooxml_cells_per_workbook=2),
        )
    assert cell_exc.value.http_status == PAYLOAD_TOO_LARGE_STATUS
    assert cell_exc.value.error_key == 'SG_IMPORT_WORKBOOK_TOO_COMPLEX'

    with pytest.raises(ShotGridDomainException) as merge_exc:
        ExcelSecurityService.validate_and_hash(
            '资产.xlsx',
            stream.getvalue(),
            ShotGridImportConfig(max_ooxml_merged_cells=2),
        )
    assert merge_exc.value.http_status == PAYLOAD_TOO_LARGE_STATUS
    assert merge_exc.value.error_key == 'SG_IMPORT_WORKBOOK_TOO_COMPLEX'


def test_ooxml_preflight_rejects_sparse_row_index_and_long_cell_text() -> None:
    namespace = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
    sparse_sheet = (
        f'<worksheet xmlns="{namespace}"><sheetData><row r="10002"><c r="A10002"/></row></sheetData></worksheet>'
    ).encode()
    sparse_contents = _replace_zip_entries(_xlsx_bytes(), {'xl/worksheets/sheet1.xml': sparse_sheet})
    with pytest.raises(ShotGridDomainException) as row_exc:
        ExcelSecurityService.validate_and_hash('镜头.xlsx', sparse_contents)
    assert row_exc.value.error_key == 'SG_IMPORT_ROW_LIMIT_EXCEEDED'

    workbook = Workbook()
    workbook.active['A1'] = '五个字符文本'
    stream = io.BytesIO()
    workbook.save(stream)
    workbook.close()
    with pytest.raises(ShotGridDomainException) as text_exc:
        ExcelSecurityService.validate_and_hash(
            '镜头.xlsx',
            stream.getvalue(),
            ShotGridImportConfig(max_cell_text_length=4),
        )
    assert text_exc.value.http_status == UNPROCESSABLE_STATUS
    assert text_exc.value.error_key == 'SG_IMPORT_CELL_TEXT_TOO_LONG'


def test_ooxml_preflight_limits_expanded_shared_strings() -> None:
    namespace = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
    shared_strings = f'<sst xmlns="{namespace}"><si><t>四字文本</t></si></sst>'.encode()
    worksheet = (
        f'<worksheet xmlns="{namespace}"><sheetData><row r="1">'
        '<c r="A1" t="s"><v>0</v></c><c r="B1" t="s"><v>0</v></c>'
        '</row></sheetData></worksheet>'
    ).encode()
    contents = _replace_zip_entries(
        _xlsx_bytes(),
        {'xl/sharedStrings.xml': shared_strings, 'xl/worksheets/sheet1.xml': worksheet},
    )

    with pytest.raises(ShotGridDomainException) as exc_info:
        ExcelSecurityService.validate_and_hash(
            '镜头.xlsx',
            contents,
            ShotGridImportConfig(max_ooxml_text_characters=10),
        )

    assert exc_info.value.http_status == PAYLOAD_TOO_LARGE_STATUS
    assert exc_info.value.error_key == 'SG_IMPORT_WORKBOOK_TOO_COMPLEX'


def test_ooxml_preflight_rejects_dtd_and_entities() -> None:
    namespace = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
    unsafe_sheet = (
        f'<!DOCTYPE worksheet [<!ENTITY x "boom">]><worksheet xmlns="{namespace}"><sheetData/></worksheet>'
    ).encode()
    contents = _replace_zip_entries(_xlsx_bytes(), {'xl/worksheets/sheet1.xml': unsafe_sheet})

    with pytest.raises(ShotGridDomainException) as exc_info:
        ExcelSecurityService.validate_and_hash('镜头.xlsx', contents)

    assert exc_info.value.http_status == UNPROCESSABLE_STATUS
    assert exc_info.value.error_key == 'SG_IMPORT_ARCHIVE_UNSAFE'


@pytest.mark.parametrize('entry_name', ['../evil.xml', '/absolute.xml', 'C:/evil.xml', 'xl\\evil.xml'])
def test_validate_and_hash_rejects_zip_path_traversal(entry_name: str) -> None:
    with pytest.raises(ShotGridDomainException) as exc_info:
        ExcelSecurityService.validate_and_hash('镜头.xlsx', _append_zip_entry(_xlsx_bytes(), entry_name))
    assert exc_info.value.error_key == 'SG_IMPORT_ARCHIVE_UNSAFE'


def test_validate_and_hash_rejects_external_links() -> None:
    with pytest.raises(ShotGridDomainException) as exc_info:
        ExcelSecurityService.validate_and_hash(
            '镜头.xlsx',
            _append_zip_entry(_xlsx_bytes(), 'xl/externalLinks/externalLink1.xml'),
        )
    assert exc_info.value.error_key == 'SG_IMPORT_EXTERNAL_LINK_NOT_ALLOWED'


def test_import_config_reads_scoped_environment(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setenv('SHOT_GRID_IMPORT_PREVIEW_TTL_SECONDS', str(TEST_PREVIEW_TTL_SECONDS))
    assert ShotGridImportConfig().preview_ttl_seconds == TEST_PREVIEW_TTL_SECONDS


@pytest.mark.asyncio
async def test_parse_in_thread_returns_sync_parser_result() -> None:
    result = await ExcelSecurityService.parse_in_thread(len, b'123')
    assert result == TEST_CONTENT_LENGTH
