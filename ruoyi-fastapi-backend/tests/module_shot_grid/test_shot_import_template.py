import hashlib
import re
from pathlib import Path
from urllib.parse import unquote
from zipfile import ZipFile

import pytest
from fastapi import FastAPI
from openpyxl import load_workbook

from common.aspect.interface_auth import CheckUserInterfaceAuth
from module_shot_grid.config import SHOT_GRID_IMPORT_CONFIG
from module_shot_grid.controller.shot_import_template_controller import (
    XLSX_MEDIA_TYPE,
    download_shot_import_template,
    shot_import_template_controller,
)
from module_shot_grid.exceptions import ShotGridDomainException
from module_shot_grid.service.shot_excel_parser import ShotExcelParser
from module_shot_grid.service.shot_import_template_service import ShotGridShotImportTemplateService

EXPECTED_SHA256 = 'b6f24078ca56295e9e6cce50bb3455af198dfffe5c08f8d85605a68c09439ece'
SERVICE_UNAVAILABLE_STATUS = 503
EXPECTED_HEADERS = (
    '场次',
    '镜头号',
    '时长(s)',
    '镜头缩略图',
    '制作内容描述',
    '景别',
    '机位',
    '镜头运动',
    '焦段(mm)',
    '场景',
    '台词/对白',
    '音效',
    '色调参考',
    '备注',
    '镜头状态',
)


def _template_bytes() -> bytes:
    return ShotGridShotImportTemplateService.TEMPLATE_PATH.read_bytes()


def test_packaged_template_digest_and_parser_statistics_are_frozen() -> None:
    contents = _template_bytes()
    parsed = ShotExcelParser(SHOT_GRID_IMPORT_CONFIG).parse(contents)

    assert hashlib.sha256(contents).hexdigest() == EXPECTED_SHA256
    assert ShotGridShotImportTemplateService.EXPECTED_SHA256 == EXPECTED_SHA256
    assert parsed.summary.model_dump(by_alias=True) == {
        'totalRows': 24,
        'validRows': 24,
        'warningRows': 0,
        'errorRows': 0,
        'distinctEpisodes': 2,
        'distinctScenes': 8,
        'distinctShots': 24,
    }
    assert [issue.error_key for issue in parsed.workbook_warnings] == ['SG_IMPORT_READONLY_COLUMNS_IGNORED']


def test_packaged_template_does_not_expose_local_or_network_paths() -> None:
    with ZipFile(ShotGridShotImportTemplateService.TEMPLATE_PATH) as workbook:
        xml_text = '\n'.join(
            workbook.read(name).decode('utf-8', errors='ignore')
            for name in workbook.namelist()
            if name.endswith(('.xml', '.rels'))
        )

    assert 'x15ac:absPath' not in xml_text
    assert re.search(r'file:(?:/{1,3}|\\)', xml_text, flags=re.IGNORECASE) is None
    assert re.search(r'(?<![A-Za-z])[A-Za-z]:[\\/]', xml_text) is None
    assert '\\\\' not in xml_text
    assert '制作人' not in xml_text


def test_packaged_template_uses_exact_business_headers_without_assignee_column() -> None:
    workbook = load_workbook(ShotGridShotImportTemplateService.TEMPLATE_PATH, read_only=True, data_only=False)
    try:
        visible_sheets = [sheet for sheet in workbook.worksheets if sheet.sheet_state == 'visible']
        assert visible_sheets
        assert ShotExcelParser.EXPECTED_HEADERS == EXPECTED_HEADERS
        for sheet in visible_sheets:
            assert tuple(sheet.cell(row=1, column=index).value for index in range(1, 16)) == EXPECTED_HEADERS
            assert not any(
                '制作人' in str(cell.value)
                for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row)
                for cell in row
                if cell.value is not None
            )
    finally:
        workbook.close()


@pytest.mark.asyncio
async def test_template_service_returns_only_verified_packaged_bytes() -> None:
    contents = await ShotGridShotImportTemplateService.get_template_bytes()

    assert contents == _template_bytes()


@pytest.mark.asyncio
@pytest.mark.parametrize('mode', ['missing', 'digest_mismatch'])
async def test_template_service_uses_stable_unavailable_error(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
    mode: str,
) -> None:
    template_path = tmp_path / 'shot-v2.xlsx'
    if mode == 'digest_mismatch':
        template_path.write_bytes(b'not-the-frozen-template')
    monkeypatch.setattr(ShotGridShotImportTemplateService, 'TEMPLATE_PATH', template_path)

    with pytest.raises(ShotGridDomainException) as exc_info:
        await ShotGridShotImportTemplateService.get_template_bytes()

    assert exc_info.value.http_status == SERVICE_UNAVAILABLE_STATUS
    assert exc_info.value.error_key == 'SG_IMPORT_TEMPLATE_UNAVAILABLE'


def test_template_route_permission_and_openapi_binary_contract() -> None:
    route = shot_import_template_controller.routes[0]
    permissions = [
        dependency.dependency.perm
        for dependency in route.dependencies
        if isinstance(dependency.dependency, CheckUserInterfaceAuth)
    ]
    assert route.path == '/shot-grid/imports/shots/template'
    assert route.methods == {'GET'}
    assert permissions == ['shotgrid:shot:import']

    app = FastAPI()
    app.include_router(shot_import_template_controller)
    operation = app.openapi()['paths']['/shot-grid/imports/shots/template']['get']
    assert XLSX_MEDIA_TYPE in operation['responses']['200']['content']


@pytest.mark.asyncio
async def test_template_route_returns_versioned_download_headers() -> None:
    response = await download_shot_import_template(None)  # type: ignore[arg-type]

    disposition = unquote(response.headers['Content-Disposition'])
    assert response.media_type == XLSX_MEDIA_TYPE
    assert response.headers['X-Shot-Grid-Template-Version'] == 'shot-v2'
    assert '镜头导入模板-shot-v2.xlsx' in disposition
    assert response.body == _template_bytes()
