import io
import re
import unicodedata
from decimal import Decimal, InvalidOperation
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell, MergedCell

from module_shot_grid.config import SHOT_GRID_IMPORT_CONFIG, ShotGridImportConfig
from module_shot_grid.entity.vo.import_common_vo import ImportIssueModel
from module_shot_grid.entity.vo.shot_import_vo import (
    SQL_BIGINT_MAX,
    SQL_INTEGER_MAX,
    ShotAssetRequirementPreviewModel,
    ShotExcelParseResultModel,
    ShotImportNormalizedRowModel,
    ShotImportPreviewRowModel,
    ShotImportPreviewSummaryModel,
)
from module_shot_grid.exceptions import shot_grid_error
from module_shot_grid.shot_number import format_shot_code


class ShotExcelParser:
    """当前镜头导入模板的纯解析器；任务委派不属于导入字段。"""

    ASSET_NAME_MAX_LENGTH = 200
    EXPECTED_HEADERS = (
        '场次',
        '镜头号',
        '时长(s)',
        '镜头缩略图',
        '制作内容描述',
        '景别',
        '机位',
        '镜头运动',
        '焦段(mm)',
        '场景',
        '台词/对白',
        '音效',
        '色调参考',
        '备注',
        '镜头状态',
    )
    SHEET_PATTERN = re.compile(r'^EP(\d{3,})$')
    SHOT_PATTERN = re.compile(r'^([0-9]+)$')
    SCENE_PATTERN = re.compile(r'^(\d+)场?$')
    READONLY_COLUMNS = frozenset({4, 15})

    def __init__(self, config: ShotGridImportConfig = SHOT_GRID_IMPORT_CONFIG) -> None:
        self.config = config

    def parse(self, contents: bytes) -> ShotExcelParseResultModel:
        try:
            workbook = load_workbook(io.BytesIO(contents), read_only=False, data_only=False, keep_links=False)
        except Exception as exc:
            raise shot_grid_error(422, 'SG_IMPORT_FILE_INVALID', '无法解析 XLSX 工作簿') from exc

        try:
            return self._parse_workbook(workbook)
        finally:
            workbook.close()

    def _parse_workbook(self, workbook: Any) -> ShotExcelParseResultModel:
        rows: list[ShotImportPreviewRowModel] = []
        workbook_warnings: list[ImportIssueModel] = []
        seen_episode_numbers: set[int] = set()
        visible_sheets = [sheet for sheet in workbook.worksheets if sheet.sheet_state == 'visible']
        hidden_sheets = [sheet.title for sheet in workbook.worksheets if sheet.sheet_state != 'visible']

        if not visible_sheets:
            self._raise('SG_IMPORT_WORKBOOK_EMPTY', '工作簿没有可见业务 Sheet')
        if hidden_sheets:
            workbook_warnings.append(
                ImportIssueModel(
                    errorKey='SG_IMPORT_HIDDEN_SHEETS_IGNORED',
                    message=f'已忽略隐藏 Sheet：{", ".join(hidden_sheets)}',
                )
            )

        physical_row_upper_bound = sum(max(sheet.max_row - 1, 0) for sheet in visible_sheets)
        if physical_row_upper_bound > self.config.max_rows_per_workbook:
            self._raise('SG_IMPORT_ROW_LIMIT_EXCEEDED', '工作簿数据行数超过限制')

        for sheet_index, sheet in enumerate(visible_sheets, start=1):
            episode_no = self._parse_sheet_name(sheet.title)
            if episode_no in seen_episode_numbers:
                self._raise('SG_IMPORT_EPISODE_DUPLICATE', f'Sheet {sheet.title} 与其他 Sheet 的集号重复')
            seen_episode_numbers.add(episode_no)
            self._validate_headers(sheet)
            rows.extend(self._parse_sheet_rows(sheet, episode_no, sheet_index))

        if len(rows) > self.config.max_rows_per_workbook:
            self._raise('SG_IMPORT_ROW_LIMIT_EXCEEDED', '工作簿数据行数超过限制')
        if not rows:
            self._raise('SG_IMPORT_WORKBOOK_EMPTY', '工作簿没有可导入的数据行')

        workbook_warnings.append(
            ImportIssueModel(
                errorKey='SG_IMPORT_READONLY_COLUMNS_IGNORED',
                message='镜头缩略图和镜头状态为只读列，导入时整表忽略',
            )
        )
        summary = self.build_summary(rows)
        return ShotExcelParseResultModel(summary=summary, rows=rows, workbookWarnings=workbook_warnings)

    def _parse_sheet_rows(self, sheet: Any, episode_no: int, sheet_index: int) -> list[ShotImportPreviewRowModel]:
        del sheet_index  # 保留参数，便于后续模板版本定义 Sheet 顺序策略。
        result: list[ShotImportPreviewRowModel] = []
        seen_shot_numbers: dict[tuple[int, int], int] = {}
        sort_ordinal = 0

        for row_number in range(2, sheet.max_row + 1):
            cells = [sheet.cell(row=row_number, column=column) for column in range(1, 16)]
            if not any(
                self._cell_has_value(cell)
                for column, cell in enumerate(cells, start=1)
                if column not in self.READONLY_COLUMNS
            ):
                continue
            sort_ordinal += 1
            row = self._parse_row(sheet.title, row_number, episode_no, sort_ordinal, cells)
            if row.normalized is not None:
                shot_key = (row.normalized.scene_no, row.normalized.shot_no)
                previous_row = seen_shot_numbers.get(shot_key)
                if previous_row is not None:
                    row.errors.append(
                        self._issue(
                            'SG_SHOT_NO_CONFLICT',
                            f'镜头号与第 {previous_row} 行在同一场次内重复；镜头号在场内唯一',
                            'shotNo',
                            sheet.title,
                            row_number,
                        )
                    )
                    row.can_import = False
                else:
                    seen_shot_numbers[shot_key] = row_number
            result.append(row)

        return result

    def _parse_row(
        self,
        sheet_name: str,
        row_number: int,
        episode_no: int,
        sort_ordinal: int,
        cells: list[Cell | MergedCell],
    ) -> ShotImportPreviewRowModel:
        errors: list[ImportIssueModel] = []

        for column, cell in enumerate(cells, start=1):
            if column in self.READONLY_COLUMNS:
                continue
            if isinstance(cell, Cell) and cell.data_type == 'f':
                errors.append(
                    self._issue(
                        'SG_IMPORT_FORMULA_NOT_ALLOWED',
                        '主数据区不允许公式单元格',
                        self._field_for_column(column),
                        sheet_name,
                        row_number,
                    )
                )

        scene = self._parse_scene(cells[0].value, sheet_name, row_number, errors)
        shot_no = self._parse_shot_no(cells[1].value, sheet_name, row_number, errors)
        duration_ms = self._parse_duration(cells[2].value, sheet_name, row_number, errors)
        description = self._required_text(cells[4].value, 'description', sheet_name, row_number, errors)
        environment_name = self._optional_text(cells[9].value)

        normalized: ShotImportNormalizedRowModel | None = None
        if scene is not None and shot_no is not None and duration_ms is not None and description is not None:
            scene_no, scene_name = scene
            requirements: list[ShotAssetRequirementPreviewModel] = []
            if environment_name:
                normalized_name = self.normalize_match_key(environment_name)
                if (
                    len(environment_name) > self.ASSET_NAME_MAX_LENGTH
                    or len(normalized_name) > self.ASSET_NAME_MAX_LENGTH
                ):
                    errors.append(
                        self._issue(
                            'SG_IMPORT_FIELD_TOO_LONG',
                            '场景名称长度不能超过 200 个字符',
                            'environmentAssetNames',
                            sheet_name,
                            row_number,
                        )
                    )
                else:
                    requirements.append(
                        ShotAssetRequirementPreviewModel(
                            rawName=environment_name,
                            normalizedName=normalized_name,
                        )
                    )

            normalized = ShotImportNormalizedRowModel(
                episodeNo=episode_no,
                episodeCode=f'EP{episode_no:03d}',
                sceneNo=scene_no,
                sceneCode=f'{scene_no:03d}',
                sceneName=scene_name,
                sortOrder=sort_ordinal * 10,
                shotNo=shot_no,
                shotCode=format_shot_code(shot_no),
                durationMs=duration_ms,
                description=description,
                shotSize=self._bounded_optional(cells[5].value, 40, 'shotSize', sheet_name, row_number, errors),
                cameraPosition=self._bounded_optional(
                    cells[6].value, 100, 'cameraPosition', sheet_name, row_number, errors
                ),
                cameraMovement=self._bounded_optional(
                    cells[7].value, 100, 'cameraMovement', sheet_name, row_number, errors
                ),
                focalLength=self._bounded_optional(
                    self._canonical_number_text(cells[8].value), 50, 'focalLength', sheet_name, row_number, errors
                ),
                assetRequirements=requirements,
                dialogue=self._optional_text(cells[10].value),
                soundEffect=self._optional_text(cells[11].value),
                colorReference=self._optional_text(cells[12].value),
                remark=self._bounded_optional(cells[13].value, 500, 'remark', sheet_name, row_number, errors),
            )

        return ShotImportPreviewRowModel(
            sheetName=sheet_name,
            rowNumber=row_number,
            normalized=normalized,
            warnings=[],
            errors=errors,
            canImport=not errors and normalized is not None,
        )

    def _validate_headers(self, sheet: Any) -> None:
        headers: list[str] = []
        for column in range(1, sheet.max_column + 1):
            value = self._optional_text(sheet.cell(row=1, column=column).value)
            if value is None:
                break
            headers.append(value)
        if tuple(headers) != self.EXPECTED_HEADERS:
            self._raise(
                'SG_IMPORT_HEADER_MISMATCH',
                f'Sheet {sheet.title} 的 A:O 表头与当前镜头模板不一致',
            )

    def _parse_sheet_name(self, sheet_name: str) -> int:
        match = self.SHEET_PATTERN.fullmatch(sheet_name)
        if not match:
            self._raise('SG_IMPORT_SHEET_NAME_INVALID', f'Sheet 名 {sheet_name} 必须使用 EP 加至少三位数字')
        try:
            episode_no = int(match.group(1))
        except ValueError:
            self._raise('SG_IMPORT_SHEET_NAME_INVALID', f'Sheet 名 {sheet_name} 的集号超出范围')
        if episode_no <= 0 or episode_no > SQL_INTEGER_MAX:
            self._raise('SG_IMPORT_SHEET_NAME_INVALID', f'Sheet 名 {sheet_name} 的集号必须在数据库整数范围内')
        return episode_no

    def _parse_scene(
        self,
        value: Any,
        sheet_name: str,
        row_number: int,
        errors: list[ImportIssueModel],
    ) -> tuple[int, str | None] | None:
        text = self._optional_text(value)
        if text == '序':
            return 0, '序'
        match = self.SCENE_PATTERN.fullmatch(text or '')
        if match:
            try:
                scene_no = int(match.group(1))
            except ValueError:
                scene_no = 0
            if 0 < scene_no <= SQL_INTEGER_MAX:
                return scene_no, None
        errors.append(
            self._issue(
                'SG_IMPORT_SCENE_INVALID',
                '场次必须为“序”或正整数编号，例如 01场、1场、001',
                'sceneNo',
                sheet_name,
                row_number,
            )
        )
        return None

    def _parse_shot_no(
        self,
        value: Any,
        sheet_name: str,
        row_number: int,
        errors: list[ImportIssueModel],
    ) -> int | None:
        text = self._optional_text(value)
        match = self.SHOT_PATTERN.fullmatch(text or '')
        if match:
            try:
                shot_no = int(match.group(1))
            except ValueError:
                shot_no = 0
            if 0 < shot_no <= SQL_INTEGER_MAX:
                return shot_no
        errors.append(
            self._issue(
                'SG_IMPORT_SHOT_NO_INVALID',
                '镜头号必须是正整数，例如 0001、0002；不再使用 S 前缀',
                'shotNo',
                sheet_name,
                row_number,
            )
        )
        return None

    def _parse_duration(
        self,
        value: Any,
        sheet_name: str,
        row_number: int,
        errors: list[ImportIssueModel],
    ) -> int | None:
        try:
            decimal_value = Decimal(str(value).strip())
            if not decimal_value.is_finite() or decimal_value < 0:
                raise InvalidOperation
            milliseconds = decimal_value * 1000
            if milliseconds != milliseconds.to_integral_value() or milliseconds > SQL_BIGINT_MAX:
                raise InvalidOperation
            return int(milliseconds)
        except (InvalidOperation, AttributeError, TypeError, ValueError):
            errors.append(
                self._issue(
                    'SG_IMPORT_DURATION_INVALID',
                    '时长必须是非负且最多三位小数的秒值',
                    'durationMs',
                    sheet_name,
                    row_number,
                )
            )
            return None

    @classmethod
    def normalize_match_key(cls, value: str) -> str:
        normalized = unicodedata.normalize('NFKC', value)
        return ' '.join(normalized.split()).casefold()

    @staticmethod
    def _canonical_number_text(value: Any) -> str | None:
        if value is None:
            return None
        if isinstance(value, bool):
            return str(value)
        if isinstance(value, (int, float, Decimal)):
            decimal_value = Decimal(str(value))
            return format(decimal_value.normalize(), 'f')
        return str(value)

    def _required_text(
        self,
        value: Any,
        field_name: str,
        sheet_name: str,
        row_number: int,
        errors: list[ImportIssueModel],
    ) -> str | None:
        text = self._optional_text(value)
        if text is None:
            errors.append(
                self._issue(
                    'SG_IMPORT_REQUIRED_FIELD_MISSING',
                    '制作内容描述不能为空',
                    field_name,
                    sheet_name,
                    row_number,
                )
            )
        return text

    def _bounded_optional(
        self,
        value: Any,
        max_length: int,
        field_name: str,
        sheet_name: str,
        row_number: int,
        errors: list[ImportIssueModel],
    ) -> str | None:
        text = self._optional_text(value)
        if text is not None and len(text) > max_length:
            errors.append(
                self._issue(
                    'SG_IMPORT_FIELD_TOO_LONG',
                    f'字段长度不能超过 {max_length} 个字符',
                    field_name,
                    sheet_name,
                    row_number,
                )
            )
            return None
        return text

    @staticmethod
    def _optional_text(value: Any) -> str | None:
        if value is None:
            return None
        text = str(value).strip()
        return text or None

    @staticmethod
    def _cell_has_value(cell: Cell | MergedCell) -> bool:
        return cell.value is not None and str(cell.value).strip() != ''

    @staticmethod
    def _field_for_column(column: int) -> str:
        fields = (
            'sceneNo',
            'shotNo',
            'durationMs',
            'thumbnail',
            'description',
            'shotSize',
            'cameraPosition',
            'cameraMovement',
            'focalLength',
            'environmentAssetNames',
            'dialogue',
            'soundEffect',
            'colorReference',
            'remark',
            'status',
        )
        return fields[column - 1]

    @staticmethod
    def _issue(
        error_key: str,
        message: str,
        field_name: str,
        sheet_name: str,
        row_number: int,
    ) -> ImportIssueModel:
        return ImportIssueModel(
            errorKey=error_key,
            message=message,
            fieldName=field_name,
            sheetName=sheet_name,
            rowNumber=row_number,
        )

    @staticmethod
    def build_summary(rows: list[ShotImportPreviewRowModel]) -> ShotImportPreviewSummaryModel:
        structural_rows = [row.normalized for row in rows if row.normalized is not None]
        valid_rows = [row for row in rows if row.can_import and row.normalized is not None]
        return ShotImportPreviewSummaryModel(
            totalRows=len(rows),
            validRows=len(valid_rows),
            warningRows=sum(bool(row.warnings) for row in rows),
            errorRows=sum(bool(row.errors) for row in rows),
            distinctEpisodes=len({row.episode_no for row in structural_rows}),
            distinctScenes=len({(row.episode_no, row.scene_no) for row in structural_rows}),
            distinctShots=len({(row.episode_no, row.scene_no, row.shot_no) for row in structural_rows}),
        )

    @staticmethod
    def _raise(error_key: str, message: str) -> Any:
        raise shot_grid_error(422, error_key, message)
