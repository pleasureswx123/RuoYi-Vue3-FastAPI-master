import hashlib
import io
import re
import unicodedata
from dataclasses import dataclass, field

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.worksheet.worksheet import Worksheet

from module_shot_grid.config import SHOT_GRID_IMPORT_CONFIG, ShotGridImportConfig
from module_shot_grid.entity.vo.asset_import_vo import (
    AssetExcelParseResultModel,
    AssetImportNormalizedRowModel,
    AssetImportPreviewRowModel,
    AssetImportPreviewSummaryModel,
    AssetImportTypeSummaryModel,
)
from module_shot_grid.entity.vo.import_common_vo import ImportIssueModel
from module_shot_grid.exceptions import shot_grid_error

_WHITESPACE_RE = re.compile(r'\s+', flags=re.UNICODE)
MAX_ASSET_NAME_LENGTH = 200
MAX_PRODUCTION_ITEM_LENGTH = 240
MAX_REMARK_LENGTH = 500
MIN_CONFLICTING_VALUES = 2
DATA_START_ROW = 2


@dataclass(slots=True)
class AssetExcelParseResult:
    """资产工作簿解析结果。"""

    rows: list[AssetImportPreviewRowModel] = field(default_factory=list)
    workbook_warnings: list[ImportIssueModel] = field(default_factory=list)
    workbook_errors: list[ImportIssueModel] = field(default_factory=list)


class AssetExcelParser:
    """只负责资产 Excel 的确定性结构解析，不执行数据库查询。"""

    HEADER_ALIASES = {
        '类型': 'asset_type',
        '资产类型': 'asset_type',
        '名称': 'asset_name',
        '资产名称': 'asset_name',
        '描述': 'asset_description',
        '分项补充要求': 'item_description',
        '制作分项描述': 'item_description',
        '资产描述': 'asset_description',
        '制作分项': 'production_item',
        '备注': 'remark',
        '状态': 'ignored_status',
        '缩略图': 'ignored_thumbnail',
        '资产缩略图': 'ignored_thumbnail',
        '最新版本': 'ignored_latest_version',
        '完成度': 'ignored_progress',
    }
    TYPE_ALIASES = {
        'character': 'Character',
        '角色': 'Character',
        'environment': 'Environment',
        '场景': 'Environment',
        'prop': 'Prop',
        '道具': 'Prop',
    }
    MERGED_PARENT_FIELDS = {'asset_type', 'asset_name', 'asset_description'}

    def __init__(self, config: ShotGridImportConfig = SHOT_GRID_IMPORT_CONFIG) -> None:
        self.config = config

    def parse(self, contents: bytes, *, file_sha256: str | None = None) -> AssetExcelParseResultModel:
        """解析 `.xlsx` 字节并返回逐行规范化结果。"""
        cls = type(self)
        file_sha256 = file_sha256 or hashlib.sha256(contents).hexdigest()
        try:
            workbook = load_workbook(
                io.BytesIO(contents),
                read_only=False,
                data_only=False,
                keep_links=False,
            )
        except Exception as exc:
            raise shot_grid_error(422, 'SG_IMPORT_FILE_INVALID', '无法解析资产 XLSX 工作簿') from exc
        result = AssetExcelParseResult()
        try:
            hidden_sheets = [sheet.title for sheet in workbook.worksheets if sheet.sheet_state != 'visible']
            if hidden_sheets:
                result.workbook_warnings.append(
                    cls._issue(
                        'SG_IMPORT_HIDDEN_SHEETS_IGNORED',
                        f'已忽略{len(hidden_sheets)}个隐藏辅助 Sheet',
                    )
                )
            visible_sheets = [sheet for sheet in workbook.worksheets if sheet.sheet_state == 'visible']
            if not visible_sheets:
                result.workbook_errors.append(cls._issue('SG_IMPORT_TEMPLATE_INVALID', '工作簿没有可见业务 Sheet'))
                return result

            total_rows = 0
            for worksheet in visible_sheets:
                sheet_rows = cls._parse_sheet(
                    worksheet,
                    file_sha256,
                    result,
                    row_limit=self.config.max_rows_per_workbook - total_rows,
                )
                total_rows += len(sheet_rows)
                result.rows.extend(sheet_rows)
                if total_rows > self.config.max_rows_per_workbook:
                    result.workbook_errors.append(
                        cls._issue(
                            'SG_IMPORT_ROW_LIMIT_EXCEEDED',
                            f'工作簿数据行不能超过{self.config.max_rows_per_workbook}行',
                        )
                    )
                    break

            if not result.rows and not result.workbook_errors:
                result.workbook_errors.append(cls._issue('SG_IMPORT_TEMPLATE_INVALID', '工作簿没有可导入的资产明细行'))
            cls._validate_duplicate_production_items(result.rows)
            cls._validate_parent_consistency(result.rows)
            if result.workbook_errors:
                first = result.workbook_errors[0]
                raise shot_grid_error(
                    422,
                    first.error_key,
                    first.message,
                    details={
                        'issues': [issue.model_dump(mode='json', by_alias=True) for issue in result.workbook_errors]
                    },
                )
            return AssetExcelParseResultModel(
                summary=cls.build_summary(result.rows),
                rows=result.rows,
                workbookWarnings=result.workbook_warnings,
            )
        finally:
            workbook.close()

    @classmethod
    def normalize_display_text(cls, value: object) -> str | None:
        """统一全半角和空白，同时保留用于展示的大小写。"""
        if value is None:
            return None
        text = unicodedata.normalize('NFKC', str(value))
        text = _WHITESPACE_RE.sub(' ', text).strip()
        return text or None

    @classmethod
    def normalize_match_key(cls, value: object) -> str | None:
        """生成资产和待匹配需求共用的精确匹配键。"""
        normalized = cls.normalize_display_text(value)
        return normalized.casefold() if normalized else None

    @classmethod
    def _parse_sheet(
        cls,
        worksheet: Worksheet,
        file_sha256: str,
        result: AssetExcelParseResult,
        *,
        row_limit: int,
    ) -> list[AssetImportPreviewRowModel]:
        header_mapping = cls._read_header(worksheet, result)
        if not header_mapping:
            return []
        if max(worksheet.max_row - 1, 0) > row_limit:
            raise shot_grid_error(422, 'SG_IMPORT_ROW_LIMIT_EXCEEDED', '工作簿数据行数超过限制')

        merge_lookup = cls._build_merge_lookup(worksheet, header_mapping)
        rows: list[AssetImportPreviewRowModel] = []
        for row_number in range(2, worksheet.max_row + 1):
            if not cls._is_effective_row(worksheet, row_number, header_mapping):
                continue
            if len(rows) >= row_limit:
                raise shot_grid_error(422, 'SG_IMPORT_ROW_LIMIT_EXCEEDED', '工作簿数据行数超过限制')
            row = cls._parse_row(worksheet, row_number, header_mapping, merge_lookup, file_sha256)
            rows.append(row)
        return rows

    @classmethod
    def _read_header(
        cls,
        worksheet: Worksheet,
        result: AssetExcelParseResult,
    ) -> dict[str, int]:
        mapping: dict[str, int] = {}
        for column in range(1, worksheet.max_column + 1):
            raw_header = cls.normalize_display_text(worksheet.cell(1, column).value)
            if raw_header is None:
                break
            field_name = cls.HEADER_ALIASES.get(raw_header)
            if field_name is None:
                result.workbook_errors.append(
                    cls._issue(
                        'SG_IMPORT_HEADER_INVALID',
                        f'Sheet {worksheet.title} 包含不支持的表头“{raw_header}”',
                        sheet_name=worksheet.title,
                    )
                )
                continue
            if field_name in mapping:
                result.workbook_errors.append(
                    cls._issue(
                        'SG_IMPORT_HEADER_DUPLICATE',
                        f'Sheet {worksheet.title} 的字段“{raw_header}”重复',
                        field_name=field_name,
                        sheet_name=worksheet.title,
                    )
                )
                continue
            mapping[field_name] = column

        for required in ('asset_type', 'asset_name'):
            if required not in mapping:
                result.workbook_errors.append(
                    cls._issue(
                        'SG_IMPORT_HEADER_REQUIRED',
                        f'Sheet {worksheet.title} 缺少必需表头',
                        field_name=required,
                        sheet_name=worksheet.title,
                    )
                )
        if any(field_name.startswith('ignored_') for field_name in mapping) and not any(
            issue.error_key == 'SG_IMPORT_READONLY_COLUMNS_IGNORED' for issue in result.workbook_warnings
        ):
            result.workbook_warnings.append(
                cls._issue(
                    'SG_IMPORT_READONLY_COLUMNS_IGNORED',
                    '状态、缩略图、最新版本和完成度等只读列在导入时忽略',
                    field_name='readonlyColumns',
                    sheet_name=worksheet.title,
                )
            )
        if result.workbook_errors:
            return {}
        return mapping

    @classmethod
    def _build_merge_lookup(
        cls,
        worksheet: Worksheet,
        header_mapping: dict[str, int],
    ) -> dict[tuple[int, int], tuple[int, int]]:
        inheritable_columns = {
            column for field_name, column in header_mapping.items() if field_name in cls.MERGED_PARENT_FIELDS
        }
        lookup: dict[tuple[int, int], tuple[int, int]] = {}
        for merged_range in worksheet.merged_cells.ranges:
            parsed_columns = set(header_mapping.values())
            intersects_data = any(merged_range.min_col <= column <= merged_range.max_col for column in parsed_columns)
            if not intersects_data:
                continue
            if (
                merged_range.min_row < DATA_START_ROW
                or merged_range.min_col != merged_range.max_col
                or merged_range.min_col not in inheritable_columns
            ):
                raise shot_grid_error(
                    422,
                    'SG_IMPORT_TEMPLATE_INVALID',
                    f'Sheet {worksheet.title} 只允许类型、名称、描述列使用单列纵向合并',
                )
            column = merged_range.min_col
            max_row = min(merged_range.max_row, worksheet.max_row)
            for row_number in range(merged_range.min_row, max_row + 1):
                lookup[(row_number, column)] = (merged_range.min_row, column)
        return lookup

    @classmethod
    def _is_effective_row(cls, worksheet: Worksheet, row_number: int, header_mapping: dict[str, int]) -> bool:
        for field_name, column in header_mapping.items():
            if field_name.startswith('ignored_'):
                continue
            cell = worksheet.cell(row_number, column)
            if isinstance(cell, MergedCell):
                continue
            if cls.normalize_display_text(cell.value) is not None:
                return True
        return False

    @classmethod
    def _parse_row(
        cls,
        worksheet: Worksheet,
        row_number: int,
        header_mapping: dict[str, int],
        merge_lookup: dict[tuple[int, int], tuple[int, int]],
        file_sha256: str,
    ) -> AssetImportPreviewRowModel:
        warnings: list[ImportIssueModel] = []
        errors: list[ImportIssueModel] = []
        raw: dict[str, object] = {}
        for field_name, column in header_mapping.items():
            if field_name.startswith('ignored_'):
                continue
            source_row, source_column = merge_lookup.get((row_number, column), (row_number, column))
            cell = worksheet.cell(source_row, source_column)
            if cell.data_type == 'f':
                errors.append(
                    cls._issue(
                        'SG_IMPORT_FORMULA_NOT_ALLOWED',
                        '导入字段不允许使用公式',
                        field_name=field_name,
                        sheet_name=worksheet.title,
                        row_number=row_number,
                    )
                )
                raw[field_name] = None
            else:
                raw[field_name] = cell.value

        type_text = cls.normalize_display_text(raw.get('asset_type'))
        asset_type = cls.TYPE_ALIASES.get(type_text.casefold()) if type_text else None
        if asset_type is None:
            errors.append(
                cls._issue(
                    'SG_ASSET_TYPE_INVALID',
                    '资产类型只允许 Character、Environment、Prop 或角色、场景、道具',
                    field_name='assetType',
                    sheet_name=worksheet.title,
                    row_number=row_number,
                )
            )

        asset_name = cls.normalize_display_text(raw.get('asset_name'))
        asset_name_key = cls.normalize_match_key(asset_name)
        if asset_name is None:
            errors.append(
                cls._issue(
                    'SG_ASSET_NAME_REQUIRED',
                    '资产名称不能为空',
                    field_name='assetName',
                    sheet_name=worksheet.title,
                    row_number=row_number,
                )
            )
        elif len(asset_name) > MAX_ASSET_NAME_LENGTH:
            errors.append(
                cls._issue(
                    'SG_IMPORT_FIELD_TOO_LONG',
                    '资产名称不能超过200个字符',
                    field_name='assetName',
                    sheet_name=worksheet.title,
                    row_number=row_number,
                )
            )
        if asset_name_key and len(asset_name_key) > MAX_ASSET_NAME_LENGTH:
            errors.append(
                cls._issue(
                    'SG_IMPORT_FIELD_TOO_LONG',
                    '资产名称规范化后不能超过200个字符',
                    field_name='assetName',
                    sheet_name=worksheet.title,
                    row_number=row_number,
                )
            )

        production_item = cls.normalize_display_text(raw.get('production_item'))
        production_item_key = cls.normalize_match_key(production_item)
        if production_item is None:
            warnings.append(
                cls._issue(
                    'SG_ASSET_PRODUCTION_ITEM_MISSING',
                    '制作分项为空，允许导入后补充，但提交图片版本前必须填写',
                    field_name='productionItem',
                    sheet_name=worksheet.title,
                    row_number=row_number,
                )
            )
        elif len(production_item) > MAX_PRODUCTION_ITEM_LENGTH:
            errors.append(
                cls._issue(
                    'SG_IMPORT_FIELD_TOO_LONG',
                    '制作分项不能超过240个字符',
                    field_name='productionItem',
                    sheet_name=worksheet.title,
                    row_number=row_number,
                )
            )
        if production_item_key and len(production_item_key) > MAX_PRODUCTION_ITEM_LENGTH:
            errors.append(
                cls._issue(
                    'SG_IMPORT_FIELD_TOO_LONG',
                    '制作分项规范化后不能超过240个字符',
                    field_name='productionItem',
                    sheet_name=worksheet.title,
                    row_number=row_number,
                )
            )

        remark = cls.normalize_display_text(raw.get('remark'))
        if remark and len(remark) > MAX_REMARK_LENGTH:
            errors.append(
                cls._issue(
                    'SG_IMPORT_FIELD_TOO_LONG',
                    '备注不能超过500个字符',
                    field_name='remark',
                    sheet_name=worksheet.title,
                    row_number=row_number,
                )
            )

        group_material = f'{worksheet.title}\0{asset_type or ""}\0{asset_name_key or ""}'
        group_key = hashlib.sha256(group_material.encode()).hexdigest()
        row_material = f'{file_sha256}\0{worksheet.title}\0{row_number}'
        import_row_key = hashlib.sha256(row_material.encode()).hexdigest()
        normalized = AssetImportNormalizedRowModel(
            assetType=asset_type,
            assetName=asset_name,
            assetNameKey=asset_name_key,
            assetGroupKey=group_key,
            assetDescription=cls.normalize_display_text(raw.get('asset_description')),
            productionItem=production_item,
            productionItemKey=production_item_key,
            itemDescription=cls.normalize_display_text(raw.get('item_description')),
            remark=remark,
            importRowKey=import_row_key,
        )
        return AssetImportPreviewRowModel(
            sheetName=worksheet.title,
            rowNumber=row_number,
            rowKey=import_row_key,
            normalized=normalized,
            warnings=warnings,
            errors=errors,
            canImport=not errors,
        )

    @classmethod
    def _validate_duplicate_production_items(cls, rows: list[AssetImportPreviewRowModel]) -> None:
        by_identity: dict[tuple[str, str, str], list[AssetImportPreviewRowModel]] = {}
        for row in rows:
            normalized = row.normalized
            if normalized.asset_type and normalized.asset_name_key and normalized.production_item_key:
                by_identity.setdefault(
                    (normalized.asset_type, normalized.asset_name_key, normalized.production_item_key),
                    [],
                ).append(row)
        for duplicate_rows in by_identity.values():
            if len(duplicate_rows) < MIN_CONFLICTING_VALUES:
                continue
            for row in duplicate_rows:
                row.errors.append(
                    cls._issue(
                        'SG_ASSET_PRODUCTION_ITEM_CONFLICT',
                        '同一资产内制作分项名称重复',
                        field_name='productionItem',
                        sheet_name=row.sheet_name,
                        row_number=row.row_number,
                    )
                )
                row.refresh_can_import()

    @classmethod
    def _validate_parent_consistency(cls, rows: list[AssetImportPreviewRowModel]) -> None:
        """同一规范资产不能在一个工作簿中声明不同的资产主描述。"""
        by_parent: dict[tuple[str, str], list[AssetImportPreviewRowModel]] = {}
        for row in rows:
            normalized = row.normalized
            if normalized.asset_type and normalized.asset_name_key:
                by_parent.setdefault((normalized.asset_type, normalized.asset_name_key), []).append(row)
        for parent_rows in by_parent.values():
            descriptions = {
                cls.normalize_display_text(row.normalized.asset_description)
                for row in parent_rows
                if row.normalized.asset_description is not None
            }
            if len(descriptions) < MIN_CONFLICTING_VALUES:
                continue
            for row in parent_rows:
                row.errors.append(
                    cls._issue(
                        'SG_ASSET_NAME_CONFLICT',
                        '同一资产在工作簿中声明了不同的资产描述',
                        field_name='assetDescription',
                        sheet_name=row.sheet_name,
                        row_number=row.row_number,
                    )
                )
                row.refresh_can_import()

    @staticmethod
    def build_summary(
        rows: list[AssetImportPreviewRowModel],
        *,
        estimated_auto_matches: int = 0,
    ) -> AssetImportPreviewSummaryModel:
        valid_rows = [row for row in rows if row.can_import]
        asset_keys = {
            (row.normalized.asset_type, row.normalized.asset_name_key)
            for row in rows
            if row.normalized.asset_type and row.normalized.asset_name_key
        }
        by_type: dict[str, AssetImportTypeSummaryModel] = {}
        for asset_type in ('Character', 'Environment', 'Prop'):
            type_rows = [row for row in rows if row.normalized.asset_type == asset_type]
            type_asset_keys = {row.normalized.asset_name_key for row in type_rows if row.normalized.asset_name_key}
            by_type[asset_type] = AssetImportTypeSummaryModel(
                assets=len(type_asset_keys),
                items=len(type_rows),
                validRows=sum(row.can_import for row in type_rows),
                warningRows=sum(bool(row.warnings) for row in type_rows),
                errorRows=sum(bool(row.errors) for row in type_rows),
            )
        return AssetImportPreviewSummaryModel(
            totalRows=len(rows),
            validRows=len(valid_rows),
            warningRows=sum(bool(row.warnings) for row in rows),
            errorRows=sum(bool(row.errors) for row in rows),
            distinctAssets=len(asset_keys),
            distinctAssetItems=len(rows),
            byType=by_type,
            estimatedAutoMatches=estimated_auto_matches,
        )

    @staticmethod
    def _issue(
        error_key: str,
        message: str,
        *,
        field_name: str | None = None,
        sheet_name: str | None = None,
        row_number: int | None = None,
    ) -> ImportIssueModel:
        return ImportIssueModel(
            errorKey=error_key,
            message=message,
            fieldName=field_name,
            sheetName=sheet_name,
            rowNumber=row_number,
        )
